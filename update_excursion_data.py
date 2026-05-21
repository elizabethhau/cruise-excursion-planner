#!/usr/bin/env python3
"""
Update data.js excursion entries with accurate activity levels and full descriptions
from the Excel reference file. Removes _review_activity:true flags.

Run from project root:
    python3 update_excursion_data.py
"""
import openpyxl
import re

EXCEL_PATH = "reference_docs/Regent 7 Seas Shore Excursions.xlsx"
DATA_JS_PATH = "data.js"

ACTIVITY_MAP = {
    "Moderate": "moderate",
    "Light": "light",
    "Strenuous": "strenuous",
    "Seated Tour": "light",
}


def load_excel_data():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 2:
                continue
            code = row[3]
            if not code or not isinstance(code, str) or code == "Excursion Code":
                continue
            if code in data:
                continue  # keep first occurrence only
            activity_raw = row[5] or "Moderate"
            desc_raw = str(row[6]) if row[6] else ""
            data[code] = {
                "activity_level": ACTIVITY_MAP.get(activity_raw, "moderate"),
                "description": desc_raw.replace("\r", ""),
            }
    return data


def escape_for_js_string(s):
    """Escape a Python string for embedding inside a JS double-quoted string."""
    s = s.replace("\\", "\\\\")
    s = s.replace('"', '\\"')
    s = s.replace("\n", "\\n")
    return s


def update_data_js(excel_data):
    with open(DATA_JS_PATH, encoding="utf-8") as f:
        lines = f.readlines()

    output = []
    current_code = None

    for line in lines:
        code_match = re.search(r'code:"([^"]+)"', line)
        if code_match:
            current_code = code_match.group(1)

        if current_code and current_code in excel_data:
            info = excel_data[current_code]

            # Update activity_level on the code line
            if code_match:
                line = re.sub(
                    r'activity_level:"[^"]+"',
                    f'activity_level:"{info["activity_level"]}"',
                    line,
                )
                # Remove _review_activity:true (with preceding comma+space)
                line = re.sub(r",\s*_review_activity:true", "", line)

            # Replace description line — build new line directly to avoid
            # re.sub interpreting \n in the replacement as an actual newline
            if re.match(r"\s+description:", line):
                desc_escaped = escape_for_js_string(info["description"])
                indent = re.match(r"(\s+)", line).group(1)
                line = f'{indent}description:"{desc_escaped}",\n'

        output.append(line)

    with open(DATA_JS_PATH, "w", encoding="utf-8") as f:
        f.writelines(output)


def verify(excel_data):
    with open(DATA_JS_PATH, encoding="utf-8") as f:
        content = f.read()

    remaining_review = content.count("_review_activity")
    desc_count = content.count("description:")
    codes_found = len(re.findall(r'code:"[A-Z]{3}-[A-Z0-9]+"', content))

    print(f"Codes in JS:             {codes_found} (expected 88)")
    print(f"Description fields:      {desc_count} (expected 88)")
    print(f"_review_activity left:   {remaining_review} (expected 0)")

    # Spot-check a known entry
    if 'activity_level:"light"' in content and 'activity_level:"moderate"' in content:
        print("Activity levels:         ✓ multiple levels present")
    if "WLG-006" in content:
        # WLG-006 should be light (Seated Tour)
        match = re.search(r'code:"WLG-006"[^\n]*activity_level:"([^"]+)"', content)
        if match:
            print(f"WLG-006 activity_level:  {match.group(1)} (expected light)")


if __name__ == "__main__":
    print("Loading Excel data...")
    excel_data = load_excel_data()
    print(f"Loaded {len(excel_data)} excursions from Excel")

    print("Updating data.js...")
    update_data_js(excel_data)

    print("\nVerification:")
    verify(excel_data)
    print("\nDone.")
